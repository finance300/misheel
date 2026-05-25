#!/usr/bin/env python3
"""
Extract fund data from "Copy of Misheel Fund Calculation FOR APP.xlsx"
into JSON files under src/data/ so the React app can render real numbers.

Re-run after the workbook changes:
    python3 scripts/extract-fund-data.py
"""
import json
import os
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Run: pip3 install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "Copy of Misheel Fund Calculation FOR APP.xlsx"
OUT = ROOT / "src" / "data"
OUT.mkdir(parents=True, exist_ok=True)


def num(v):
    if v is None:
        return None
    try:
        f = float(v)
        return f
    except (TypeError, ValueError):
        return None


def iso(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return v


def col(ws, r, c):
    return ws.cell(r, c).value


wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------- Portfolio Builder: FX rates + equity currencies ----------
pb = wb["Portfolio Builder"]
fx_rates = {}
# FX block starts at row 4 (after header at row 3)
for r in range(4, 20):
    cur = col(pb, r, 2)
    rate = num(col(pb, r, 3))
    if not cur or rate is None:
        break
    fx_rates[str(cur)] = rate

# Build ticker -> {currency, type} map from Portfolio Builder equities (header row 12, data row 13+)
ticker_meta = {}
for r in range(13, pb.max_row + 1):
    cur = col(pb, r, 3)
    typ = col(pb, r, 4)
    ticker = col(pb, r, 5)
    if cur is None or ticker is None:
        continue
    ticker_meta[str(ticker).strip()] = {
        "currency": str(cur).strip(),
        "type": str(typ).strip() if typ else "Stock",
    }

# ---------- NAV sheet ----------
nav = wb["NAV"]
summary = {
    "totalEquities": num(col(nav, 2, 3)),
    "totalBonds": num(col(nav, 3, 3)),
    "liabilities": num(col(nav, 4, 3)),
    "netAssets": num(col(nav, 5, 3)),
    "totalUnits": num(col(nav, 6, 3)),
    "navPerUnit": num(col(nav, 7, 3)),
}

cash = []
for r in range(10, 15):
    label = col(nav, r, 2)
    amount = num(col(nav, r, 3))
    mnt_value = num(col(nav, r, 4))
    if label is None or mnt_value is None:
        continue
    label_str = str(label)
    # MNT/USD/EUR rows have FX in fx_rates; other rows (e.g. payables) are already in MNT.
    currency = label_str if label_str in fx_rates else "MNT"
    cash.append({
        "label": label_str,
        "currency": currency,
        "amount": amount if amount is not None else mnt_value,
        "mntValue": mnt_value,
    })

bonds = []
for r in range(17, 30):
    ticker = col(nav, r, 2)
    if ticker is None:
        break
    bonds.append({
        "ticker": str(ticker),
        "qty": num(col(nav, r, 3)),
        "costBasis": num(col(nav, r, 4)),
        "avgPrice": num(col(nav, r, 5)),
        "yield": num(col(nav, r, 6)),
        "initialDate": iso(col(nav, r, 7)),
        "calcDate": iso(col(nav, r, 8)),
        "interestAmount": num(col(nav, r, 9)),
        "totalValue": num(col(nav, r, 10)),
        "allocationPct": num(col(nav, r, 11)),
    })

equities = []
for r in range(24, 200):
    ticker = col(nav, r, 2)
    qty = col(nav, r, 3)
    if ticker is None and qty is None:
        break
    if ticker is None:
        continue
    ticker_str = str(ticker).strip()
    meta = ticker_meta.get(ticker_str, {"currency": "MNT", "type": "Stock"})
    equities.append({
        "ticker": ticker_str,
        "currency": meta["currency"],
        "type": meta["type"],
        "qty": num(qty),
        "costBasis": num(col(nav, r, 4)),
        "avgPrice": num(col(nav, r, 5)),
        "price": num(col(nav, r, 6)),
        "mcap": num(col(nav, r, 7)),
        "unrealizedPnL": num(col(nav, r, 8)),
        "unrealizedPnLPct": num(col(nav, r, 9)),
        "mcapMnt": num(col(nav, r, 10)),
        "allocationPct": num(col(nav, r, 11)),
    })

# ---------- Liabilities ----------
liab = wb["Liabilities"]
liabilities = {
    "totalAccrued": num(col(liab, 2, 3)),
    "totalCollected": num(col(liab, 3, 3)),
    "total": num(col(liab, 4, 3)),
}

# Per-investor subscription records (one row per subscription event).
# Has the "Буцаасан нэгж эрх" (returned units) column that the Users sheet lacks.
investor_subscriptions = []
liab_header_row = None
for r in range(1, 30):
    if col(liab, r, 1) == "№":
        liab_header_row = r
        break

if liab_header_row:
    for r in range(liab_header_row + 1, liab.max_row + 1):
        idx = col(liab, r, 1)
        if idx is None:
            continue
        investor_subscriptions.append({
            "no": num(idx),
            "lastName": col(liab, r, 2),
            "firstName": col(liab, r, 3),
            "registerNumber": col(liab, r, 4),
            "phone": col(liab, r, 5),
            "email": col(liab, r, 6),
            "unitsPurchased": num(col(liab, r, 7)),
            "unitsReturned": num(col(liab, r, 8)),
            "unitsActive": num(col(liab, r, 9)),
            "unitValue": num(col(liab, r, 10)),
            "transferAmount": num(col(liab, r, 11)),
            "transferDate": iso(col(liab, r, 12)),
            "calcDate": iso(col(liab, r, 13)),
            "daysHeld": num(col(liab, r, 14)),
            "dailyFee": num(col(liab, r, 15)),
            "accruedFee": num(col(liab, r, 16)),
            "annualFee": num(col(liab, r, 17)),
        })

with open(OUT / "fund-investor-subscriptions.json", "w", encoding="utf-8") as f:
    json.dump(investor_subscriptions, f, ensure_ascii=False, indent=2)
print(f"[investors] {len(investor_subscriptions)} subscription records")

with open(OUT / "fund-nav.json", "w", encoding="utf-8") as f:
    json.dump(
        {
            "summary": summary,
            "fxRates": fx_rates,
            "cash": cash,
            "bonds": bonds,
            "equities": equities,
            "liabilities": liabilities,
        },
        f,
        ensure_ascii=False,
        indent=2,
    )

print(f"[nav]  Net Assets: {summary['netAssets']:,.0f} MNT / per unit: {summary['navPerUnit']:.4f}")
print(f"[nav]  bonds={len(bonds)}, equities={len(equities)}, cash rows={len(cash)}")

# ---------- Trade Confirmation ----------
tc = wb["Trade Confirmation"]
trades = []
for r in range(2, tc.max_row + 1):
    if col(tc, r, 1) is None and col(tc, r, 2) is None:
        continue
    trades.append({
        "securityType": col(tc, r, 1),
        "type": col(tc, r, 2),
        "currency": col(tc, r, 3),
        "tradeDate": iso(col(tc, r, 4)),
        "settleDate": iso(col(tc, r, 5)),
        "securityName": col(tc, r, 6),
        "ticker": col(tc, r, 7),
        "instrumentType": col(tc, r, 8),
        "portfolioClass": col(tc, r, 9),
        "quantity": num(col(tc, r, 10)),
        "unitPrice": num(col(tc, r, 11)),
        "secTotal": num(col(tc, r, 12)),
        "feePerc": num(col(tc, r, 13)),
        "feeAmount": num(col(tc, r, 14)),
        "fixedFee": num(col(tc, r, 15)),
        "totalFee": num(col(tc, r, 16)),
        "total": num(col(tc, r, 17)),
        "exchangeRate": col(tc, r, 18),
        "totalUsd": col(tc, r, 19),
        "stockSplit": col(tc, r, 20),
        "description": col(tc, r, 21),
    })

with open(OUT / "fund-trades.json", "w", encoding="utf-8") as f:
    json.dump(trades, f, ensure_ascii=False, indent=2)

print(f"[trades]  {len(trades)} trade confirmations")

# ---------- Per-ticker BUY/SELL/TOTAL/TOTAL (After Split) summary ----------
# Lives in columns 23..29 of the Trade Confirmation sheet (the right-hand block).
ticker_summary = []
for r in range(2, tc.max_row + 1):
    typ = col(tc, r, 23)
    ticker = col(tc, r, 25)
    if not ticker:
        continue
    ticker_summary.append({
        "type": typ,
        "currency": col(tc, r, 24),
        "ticker": str(ticker).strip(),
        "buy": num(col(tc, r, 26)),
        "sell": num(col(tc, r, 27)),
        "total": num(col(tc, r, 28)),
        "totalAfterSplit": num(col(tc, r, 29)),
    })

with open(OUT / "fund-trade-summary.json", "w", encoding="utf-8") as f:
    json.dump(ticker_summary, f, ensure_ascii=False, indent=2)
split_count = sum(
    1 for s in ticker_summary
    if s["total"] is not None and s["totalAfterSplit"] is not None and abs(s["total"] - s["totalAfterSplit"]) > 0.001
)
print(f"[summary] {len(ticker_summary)} tickers · {split_count} with active split")

# ---------- Bank accounts ----------
bank_sheets = [
    {"name": "Bank 0134 (USD)", "accountNumber": "0134", "currency": "USD", "alias": "USD trading"},
    {"name": "Bank 8863 (MNT)", "accountNumber": "8863", "currency": "MNT", "alias": "MNT main"},
    {"name": "Bank 0055 (MNT)", "accountNumber": "0055", "currency": "MNT", "alias": "MNT secondary"},
]

bank_accounts = []
for cfg in bank_sheets:
    ws = wb[cfg["name"]]
    transactions = []
    last_balance = None
    last_date = None
    for r in range(2, ws.max_row + 1):
        date = col(ws, r, 1)
        if date is None:
            continue
        end_balance = num(col(ws, r, 5))
        debit = num(col(ws, r, 3))
        credit = num(col(ws, r, 4))
        if end_balance is None and debit is None and credit is None:
            continue
        rec = {
            "date": iso(date) if isinstance(date, (datetime, date.__class__)) else str(date) if not isinstance(date, (datetime,)) else iso(date),
            "openingBalance": num(col(ws, r, 2)),
            "debit": debit,
            "credit": credit,
            "endingBalance": end_balance,
            "description": col(ws, r, 6),
            "counterpartAccount": col(ws, r, 7),
        }
        if cfg["currency"] == "USD":
            rec["fxRate"] = num(col(ws, r, 8))
            rec["debitMnt"] = col(ws, r, 9)
            rec["creditMnt"] = col(ws, r, 10)
        transactions.append(rec)
        if end_balance is not None:
            last_balance = end_balance
            last_date = rec["date"]
    bank_accounts.append({
        **cfg,
        "currentBalance": last_balance,
        "asOf": last_date,
        "transactionCount": len(transactions),
        "transactions": transactions[-100:],  # keep most recent 100 rows
    })

with open(OUT / "fund-banks.json", "w", encoding="utf-8") as f:
    json.dump(bank_accounts, f, ensure_ascii=False, indent=2)

print(f"[banks] {len(bank_accounts)} accounts")
for a in bank_accounts:
    print(f"  {a['accountNumber']} {a['currency']:3s}  balance={a['currentBalance']}  txns={a['transactionCount']}")

print(f"\nWrote JSON to {OUT}")
